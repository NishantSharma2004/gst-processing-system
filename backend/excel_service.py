import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import io
import os
import sqlite3
from .database import get_db

GST_REGEX = re.compile(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$')

def clean_gstin(val) -> str:
    if pd.isna(val):
        return ""
    val_str = str(val).strip().upper()
    val_str = re.sub(r'\s+', '', val_str)
    return val_str

def parse_and_clean_excel(file_bytes: bytes, filename: str, selected_column: str = None):
    valid_items = {} # gstin -> list of "sheet_name:row_num"
    invalid_items = []
    total_rows = 0
    available_sheets = []

    if filename.endswith('.csv'):
        df = pd.read_csv(io.BytesIO(file_bytes))
        sheets_dict = {'Sheet1': df}
        available_sheets = ['Sheet1']
    else:
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
        available_sheets = xls.sheet_names
        sheets_dict = {sheet: pd.read_excel(xls, sheet_name=sheet) for sheet in available_sheets}

    for sheet_name, df in sheets_dict.items():
        total_rows += len(df)

        # Detect GST column for this sheet
        gst_col = selected_column
        if not gst_col or gst_col not in df.columns:
            for col in df.columns:
                col_str = str(col).lower()
                if 'gst' in col_str or 'number' in col_str or 'pin' in col_str or 'code' in col_str:
                    gst_col = col
                    break
            
            if not gst_col:
                best_col = None
                max_matches = -1
                for col in df.columns:
                    matches = sum(1 for v in df[col].dropna() if GST_REGEX.match(clean_gstin(v)))
                    if matches > max_matches:
                        max_matches = matches
                        best_col = col
                gst_col = best_col if (best_col and max_matches > 0) else df.columns[0]

        for idx, row in df.iterrows():
            row_num = idx + 2
            raw_val = row[gst_col]
            cleaned = clean_gstin(raw_val)

            if not cleaned:
                continue

            ref_str = f"{sheet_name}:R{row_num}"
            if GST_REGEX.match(cleaned):
                if cleaned not in valid_items:
                    valid_items[cleaned] = []
                valid_items[cleaned].append(ref_str)
            else:
                invalid_items.append({
                    'gstin': cleaned if cleaned else str(raw_val),
                    'sheet_name': sheet_name,
                    'row_num': str(row_num),
                    'error_type': 'Invalid GSTIN Format',
                    'error_message': 'Does not match 15-character GSTIN pattern'
                })

    unique_gst_count = len(valid_items)
    duplicates_removed = sum(len(refs) - 1 for refs in valid_items.values())
    valid_gst_count = sum(len(refs) for refs in valid_items.values())
    invalid_gst_count = len(invalid_items)

    return {
        'gst_column': 'Auto Detected Across Sheets',
        'columns': available_sheets,
        'total_rows': total_rows,
        'valid_gst_count': valid_gst_count,
        'invalid_gst_count': invalid_gst_count,
        'unique_gst_count': unique_gst_count,
        'duplicates_removed': duplicates_removed,
        'valid_items': valid_items,
        'invalid_items': invalid_items
    }

def generate_3sheet_excel(job_id: str) -> bytes:
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM processing_jobs WHERE job_id = ?", (job_id,))
    job = dict(cursor.fetchone())

    cursor.execute("SELECT * FROM processing_items WHERE job_id = ?", (job_id,))
    items = [dict(r) for r in cursor.fetchall()]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    border = Border(left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9'))

    # Group items by sheet name if available
    sheet_groups = {}
    for item in items:
        if item['status'] == 'Success':
            ref = item['original_row_numbers'] or 'Sheet1'
            sheet_name = ref.split(':')[0] if ':' in ref else 'GST Details'
            if sheet_name not in sheet_groups:
                sheet_groups[sheet_name] = []
            sheet_groups[sheet_name].append(item)

    if not sheet_groups:
        sheet_groups['GST Details'] = []

    # Create a result sheet for each original sheet
    for sheet_name, group_items in sheet_groups.items():
        title = sheet_name[:30] # Excel 31 char max limit
        ws = wb.create_sheet(title=title)
        headers = ['GST Number', 'Legal Name', 'Trade Name', 'GST Status', 'Business Type']
        ws.append(headers)

        for item in group_items:
            cursor.execute("SELECT * FROM gst_records WHERE gstin = ?", (item['gstin'],))
            rec = cursor.fetchone()
            if rec:
                ws.append([
                    rec['gstin'],
                    rec['legal_name'] or 'Not Available',
                    rec['trade_name'] or 'Not Available',
                    rec['gst_status'] or 'Not Available',
                    rec['business_type'] or 'Not Available'
                ])
            else:
                ws.append([item['gstin'], 'Not Available', 'Not Available', 'Not Available', 'Not Available'])

    # Errors Sheet
    ws_err = wb.create_sheet(title='Errors')
    ws_err.append(['Sheet Name', 'GST Number', 'Error Type', 'Error Message', 'Retry Count', 'Status'])

    for item in items:
        if item['status'] in ['Failed', 'Invalid']:
            ref = item['original_row_numbers'] or 'Sheet1'
            sheet_name = ref.split(':')[0] if ':' in ref else 'Sheet1'
            ws_err.append([
                sheet_name,
                item['gstin'],
                item['error_type'] or 'Search Error',
                item['error_message'] or 'GST details unavailable',
                item['retry_count'],
                item['status']
            ])

    # Summary Sheet
    ws_sum = wb.create_sheet(title='Summary')
    ws_sum.append(['Metric Name', 'Metric Value'])
    ws_sum.append(['Filename', job['filename']])
    ws_sum.append(['Total Input Rows', job['total_rows']])
    ws_sum.append(['Valid GST Numbers', job['valid_gst_count']])
    ws_sum.append(['Invalid GST Numbers', job['invalid_gst_count']])
    ws_sum.append(['Unique GST Numbers Searched', job['unique_gst_count']])
    ws_sum.append(['Duplicates Removed', job['duplicates_removed']])
    ws_sum.append(['Successfully Processed', job['success_count']])
    ws_sum.append(['Failed / Not Found', job['failed_count']])
    ws_sum.append(['Job Status', job['status']])
    ws_sum.append(['Created At', str(job['created_at'])])
    ws_sum.append(['Completed At', str(job['completed_at']) if job['completed_at'] else 'N/A'])

    # Formatting
    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for col_num, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
            max_len = 0
            for cell in col_cells:
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    conn.close()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
